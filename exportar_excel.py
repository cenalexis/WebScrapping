"""
exportar_excel.py  —  Exporta la BD de vacantes a Excel
========================================================
Uso:
    python exportar_excel.py
    python exportar_excel.py --desde 2026-05-01   (solo desde esa fecha)
    python exportar_excel.py --salida mi_reporte.xlsx

Genera un .xlsx con 4 hojas:
    vacantes        : todos los registros con campos legibles
    resumen         : conteos por portal y fecha
    por_ubicacion   : cuántas vacantes por ciudad/provincia
    cargos_top100   : los 100 cargos más publicados
"""

import os, sys, sqlite3, argparse
from datetime import datetime

import subprocess
subprocess.run([sys.executable, "-m", "pip", "install", "-q", "pandas", "openpyxl"], check=False)

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH    = r"C:\Users\alexis\Documents\CISE_2026\vacantes_laborales.db"
SALIDA_DIR = r"C:\Users\alexis\Documents\CISE_2026\exports"


def conectar():
    if not os.path.exists(DB_PATH):
        print(f"ERROR: No se encontró la BD en:\n  {DB_PATH}")
        print("Asegúrate de haber corrido scraper_mt_v2.py al menos una vez.")
        sys.exit(1)
    return sqlite3.connect(DB_PATH)


def cargar_vacantes(desde: str | None = None) -> pd.DataFrame:
    conn = conectar()
    filtro = "WHERE v.fecha_extraccion >= ?" if desde else ""
    params = [desde] if desde else []

    query = f"""
    SELECT
        v.id,
        p.nombre            AS portal,
        v.cargo_raw         AS cargo,
        v.empresa_raw       AS empresa,
        v.ubicacion_raw     AS ubicacion,
        v.modalidad_raw     AS modalidad,
        v.jornada_raw       AS jornada,
        v.contrato_raw      AS contrato,
        v.area_raw          AS area,
        v.subarea_raw       AS subarea,
        v.industria_raw     AS industria,
        v.vacantes_num      AS num_vacantes,
        v.salario_min,
        v.salario_max,
        CASE v.salario_a_convenir WHEN 1 THEN 'A convenir' ELSE 'Definido' END AS salario_tipo,
        v.experiencia_anos,
        v.instruccion_raw   AS nivel_instruccion,
        v.idioma_raw        AS idioma,
        v.licencia_raw      AS licencia,
        v.descripcion_raw   AS descripcion,
        v.requisitos_raw    AS requisitos,
        v.beneficios_raw    AS beneficios,
        v.url_detalle       AS url,
        substr(v.fecha_extraccion, 1, 10) AS fecha_extraccion
    FROM vacantes v
    LEFT JOIN portales p ON v.portal_id = p.id
    {filtro}
    ORDER BY v.fecha_extraccion DESC
    """
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def _estilizar(ws, color: str = "1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=color)
    bold = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        ancho = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ancho + 2, 60)
    ws.freeze_panes = "A2"


def exportar(salida: str = "", desde: str | None = None):
    print(f"Leyendo: {DB_PATH}")
    df = cargar_vacantes(desde=desde)
    print(f"  {len(df):,} vacantes")

    if df.empty:
        print("La BD está vacía. Corre primero scraper_mt_v2.py.")
        return

    os.makedirs(SALIDA_DIR, exist_ok=True)
    if not salida:
        ts     = datetime.now().strftime("%Y%m%d_%H%M")
        salida = os.path.join(SALIDA_DIR, f"vacantes_{ts}.xlsx")

    # Hojas de análisis
    resumen = (df.groupby(["portal", "fecha_extraccion"])
               .agg(total=("id","count"),
                    con_salario=("salario_min", lambda x: x.notna().sum()),
                    con_experiencia=("experiencia_anos", lambda x: x.notna().sum()))
               .reset_index().sort_values(["portal","fecha_extraccion"], ascending=False))

    por_ubic = (df.groupby("ubicacion").agg(total=("id","count"))
                .reset_index().sort_values("total", ascending=False).head(50))

    cargos = (df.groupby("cargo")
              .agg(publicaciones=("id","count"),
                   vacantes_total=("num_vacantes","sum"),
                   empresas=("empresa","nunique"))
              .reset_index().sort_values("publicaciones", ascending=False).head(100))

    cols_orden = [
        "portal","cargo","empresa","ubicacion","modalidad","jornada","contrato",
        "area","subarea","industria","num_vacantes","salario_min","salario_max",
        "salario_tipo","experiencia_anos","nivel_instruccion","idioma","licencia",
        "fecha_extraccion","url","descripcion","requisitos","beneficios",
    ]
    cols_orden = [c for c in cols_orden if c in df.columns]

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df[cols_orden].to_excel(writer, sheet_name="vacantes",      index=False)
        resumen.to_excel(        writer, sheet_name="resumen",       index=False)
        por_ubic.to_excel(       writer, sheet_name="por_ubicacion", index=False)
        cargos.to_excel(         writer, sheet_name="cargos_top100", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(salida)
    for nombre, color in [
        ("vacantes","1F4E79"), ("resumen","375623"),
        ("por_ubicacion","7B3F00"), ("cargos_top100","4B0082"),
    ]:
        if nombre in wb.sheetnames:
            _estilizar(wb[nombre], color)
    wb.save(salida)

    print(f"\nArchivo listo:\n  {salida}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Exporta vacantes a Excel")
    p.add_argument("--salida", default="")
    p.add_argument("--desde",  default=None, help="Fecha mínima YYYY-MM-DD")
    a = p.parse_args()
    exportar(a.salida, a.desde)
